import openpyxl
import openpyxl.utils.cell as ut_cell
import os

cwd = os.getcwd()
filename = "D:\\HT_24050211003_조민희\\pycharm\\output\\df.xlsx"
filepath = os.path.join(cwd, "output", filename)

wb = openpyxl.load_workbook(filepath)
ws = wb['Sheet1']

# 샐 객체
b2 = ws['B2']
print(b2, b2.coordinate, b2.column, b2.row, b2.value)
print(ut_cell.get_column_letter(b2.column))
print(ut_cell.column_index_from_string('AB'))


c4 = ws.cell(row=4,column=3)
print(c4,c4.coordinate, c4.value)

b2.value = 50
c4.value = 50
print(b2.value)
print(c4.value)

# 워크북의 변경내용을 새로운 파일에 저장
wb.save(os.path.join(cwd, "output", "df2.xlsx"))